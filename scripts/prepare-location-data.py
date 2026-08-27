#!/usr/bin/env python3
"""Normalize LGD Excel exports into CSV files for the Node database importer."""

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


DATASETS = {
    "states": {
        "pattern": "All_Stateof_India_*.xlsx",
        "columns": {
            "code": "State Code",
            "name": "State Name (In English)",
            "local_name": "State Name (In Local)",
        },
    },
    "districts": {
        "pattern": "All_Districtof_India_*.xlsx",
        "columns": {
            "state_code": "State Code",
            "code": "District Code",
            "name": "District Name(In English)",
        },
    },
    "subdistricts": {
        "pattern": "All_Sub_Districtof_India_*.xlsx",
        "columns": {
            "state_code": "State Code",
            "district_code": "District Code",
            "code": "Sub-district Code",
            "name": "Sub-district Name",
        },
    },
    "villages": {
        "pattern": "All_Villagesof_India_*.xlsx",
        "columns": {
            "state_code": "State Code",
            "district_code": "District Code",
            "subdistrict_code": "Sub-District Code",
            "code": "Village Code",
            "name": "Village Name (In English)",
            "local_name": "Village Name (In Local)",
        },
    },
}

PINCODE_COLUMNS = {
    "code": "pincode",
    "state_name": "statename",
    "district_name": "district",
}


def normalized_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def source_file(input_dir, pattern):
    files = sorted(input_dir.glob(pattern))
    if len(files) != 1:
        raise ValueError(
            f"Expected exactly one {pattern} file in {input_dir}, found {len(files)}."
        )
    return files[0]


def prepare_dataset(dataset_name, config, input_dir, output_dir):
    source = source_file(input_dir, config["pattern"])
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    next(rows, None)  # report title
    headers = [normalized_value(value) for value in next(rows, ())]
    column_indexes = {}
    for field, header in config["columns"].items():
        if header not in headers:
            raise ValueError(f"{source.name} is missing required column: {header}")
        column_indexes[field] = headers.index(header)

    output_path = output_dir / f"{dataset_name}.csv"
    count = 0
    with output_path.open("w", newline="", encoding="utf-8") as output_file:
        writer = csv.DictWriter(
            output_file,
            fieldnames=config["columns"].keys(),
            lineterminator="\n",
        )
        writer.writeheader()
        for row_number, row in enumerate(rows, start=3):
            result = {
                field: normalized_value(row[index] if index < len(row) else None)
                for field, index in column_indexes.items()
            }
            if not result["code"] or not result["name"]:
                raise ValueError(
                    f"{source.name} row {row_number} has no code or English name."
                )
            writer.writerow(result)
            count += 1

    workbook.close()
    return {"sourceFile": source.name, "rows": count}


def prepare_pincodes(input_dir, output_dir):
    source = input_dir / "pincode.csv"
    if not source.is_file():
        raise ValueError(f"Expected pincode.csv in {input_dir}.")
    output_path = output_dir / "pincodes.csv"
    count = 0
    with source.open("r", newline="", encoding="utf-8-sig") as input_file, output_path.open(
        "w", newline="", encoding="utf-8"
    ) as output_file:
        reader = csv.DictReader(input_file)
        headers = reader.fieldnames or []
        for header in PINCODE_COLUMNS.values():
            if header not in headers:
                raise ValueError(f"{source.name} is missing required column: {header}")
        writer = csv.DictWriter(
            output_file,
            fieldnames=PINCODE_COLUMNS.keys(),
            lineterminator="\n",
        )
        writer.writeheader()
        for row_number, row in enumerate(reader, start=2):
            result = {
                field: normalized_value(row.get(header))
                for field, header in PINCODE_COLUMNS.items()
            }
            if not result["code"] or not result["state_name"] or not result["district_name"]:
                raise ValueError(
                    f"{source.name} row {row_number} is missing PIN, state, or district."
                )
            writer.writerow(result)
            count += 1
    return {"sourceFile": source.name, "rows": count}


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", default="locations_data")
    parser.add_argument("--output-dir", default=".location-import")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    if not input_dir.is_dir():
        raise ValueError(f"Input directory does not exist: {input_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)

    datasets = {
        name: prepare_dataset(name, config, input_dir, output_dir)
        for name, config in DATASETS.items()
    }
    if (input_dir / "pincode.csv").is_file():
        datasets["pincodes"] = prepare_pincodes(input_dir, output_dir)
    metadata = {
        "sources": ["Local Government Directory (LGD)", "pincode.csv"],
        "preparedAt": datetime.now(timezone.utc).isoformat(),
        "datasets": datasets,
    }
    (output_dir / "metadata.json").write_text(
        json.dumps(metadata, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(metadata, indent=2))


if __name__ == "__main__":
    main()
